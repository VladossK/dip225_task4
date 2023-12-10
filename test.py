import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook 

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

name=[]
name_unc = {}
# program read information from people.csv file and put all data in name list.
with open("people.csv", "r") as f:
    next(f)
    for file in f:
        row=file.rstrip().split(",")
        fname=[row[2], row[3]]
        fname_str=" ".join(fname)
        name.append(str(fname_str))


url = "https://emn178.github.io/online-tools/crc32.html"
driver.get(url)
time.sleep(2)


namesUnc = []
for names in name:
        find1 = driver.find_element(By.ID, "input")
        last = find1.send_keys(names)    
        find = driver.find_element(By.ID, "output")
        temp = find.get_attribute("value")
        namesUnc.append(temp)
        find1.clear()
        name_unc[names] = temp

wb = load_workbook('salary.xlsx')
ws = wb.active

salary_dict = {}
for row in ws.iter_rows(values_only=True):
    coded_name, salary = row[0], row[1]
    if coded_name in namesUnc:
        if coded_name not in salary_dict:
            salary_dict[coded_name] = 0
        salary_dict[coded_name] += salary


for idx, (name, coded_name) in enumerate(name_unc .items(), start=1):
    ws.cell(row=idx, column=3, value=name)
    ws.cell(row=idx, column=4, value=salary_dict.get(coded_name, 0))
   
ws.save("darbnieku_algas.xslx")


        
                
                        
        
                

                




    
    